import os
import random
import re
import time
import traceback
from datetime import datetime
from lxml import etree
import requests
from retrying import retry
from openpyxl import Workbook, load_workbook

req_session = requests.session()
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,'
              'application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Referer': 'https://www.gxghj.cn/cles/list-141-4.html',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 '
                  'Safari/537.36',
    'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
}
req_session.headers = headers
title = "柳州"
host = "https://www.gxghj.cn"
data_list = []


def save_data(file_path: str):
    data_list.reverse()
    if not os.path.exists(file_path):
        # 文件不存在，创建新的 Excel 文件并写入数据
        workbook = Workbook()
        sheet = workbook.active
        print("数据写入新的 Excel 文件")
    else:
        # 文件存在，打开现有的 Excel 文件并追加数据
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # 写入数据
        print("数据追加到现有的 Excel 文件")
    for item in data_list:
        # 写入表头
        header = list(item.keys())
        sheet.append(header)
        # 写入数据
        values = list(item.values())
        sheet.append(values)
        sheet.append([""])
    # 保存文件
    workbook.save(file_path)


@retry(stop_max_attempt_number=5, wait_random_min=1000, wait_random_max=2000, )
def get_content(url: str) -> str:
    response = req_session.get(url)
    if "?WebShieldSessionVerify" in response.text:
        # verify_str = re.findall(r"html\?(.*?)\"", response.text)
        # url += "?" + verify_str[0]
        time.sleep(random.uniform(0, 1))
        raise Exception
    return response.text


def parse_articles_list(html_text: str) -> tuple:
    arr, max_page = [], 0
    if not html_text:
        return arr, max_page
    try:
        # 使用解析器解析 HTML 字符串
        tree = etree.HTML(html_text)
        max_page = tree.xpath('//div[@class="digg"]/*[text()="下一页»"]')[0].getprevious().text.strip()
        # 使用 XPath 表达式来提取数据
        li_elements = tree.xpath(f'//div[@class="newsList"]/ul/li/a[contains(.//div[@class="newsTitle"], "{title}")]')
        arr = [
            (host + li.xpath('@href')[0], li.xpath('.//div[@class="newsTitle"]/text()')[0])
            for li in li_elements
        ]
        print(arr)
    except Exception as e:
        print(f"------------------------ 解析文章列表异常 {e} ------------------------")
    return arr, int(max_page)


def parse_data(html_text: str) -> dict:
    data = {}
    try:
        # 使用解析器解析 HTML 字符串
        tree = etree.HTML(html_text)
        # title = tree.xpath("//div[@class='titleWrap']/div[@class='title']/text()")[0].strip()
        # matches = re.findall(r"\d+", title)
        title = tree.xpath("//div[@class='info']/text()")[0].strip()
        matches = re.findall(r"(\d+-\d+-\d+)", title)
        date_obj = datetime.strptime(matches[0], "%Y-%m-%d").date()

        # 将日期对象转换为指定格式的日期字符串
        formatted_date_str = date_obj.strftime("%Y/%m/%d")
        data["日期"] = formatted_date_str
        # 使用 XPath 表达式来提取数据
        rows = tree.xpath('//div[@class="content"]/table//tr/td')
        for row in rows:
            name = row.xpath('.//span/text()')
            if not name:
                continue
            if name[0] in ['岩滩站', '天峨站', '大化站']:
                values = row.getnext().xpath('.//span/text()')[0]
                data[name[0] + "水位"] = values.strip() + "米"
        hubs = tree.xpath('//span[contains(text(), "下泄流量")]')
        if len(hubs) == 1:
            d = re.findall(r'(\w+下泄流量)\s*([\d\.]+m³/s)', hubs[0].text)
            for i in d:
                if i[0][:3] in ["龙滩枢", "岩滩枢", "大化枢", "百龙滩", "乐滩枢", "桥巩枢", "大藤峡", "龙滩电", "岩滩电", "大化电", "乐滩电", "桥巩电"]:
                    data[i[0]] = i[1]
        for span in hubs:
            span_text = span.text[1:4]
            if span_text in ["龙滩枢", "岩滩枢", "大化枢", "百龙滩", "乐滩枢", "桥巩枢", "大藤峡", "龙滩电", "岩滩电", "大化电", "乐滩电", "桥巩电"]:
                water_flow = span.xpath('following-sibling::span[1]//text()')[0]
                data[span.text[1:]] = water_flow
    except Exception as e:
        traceback.print_exc()
        print(f"------------------------ 解析水位数据异常 {e} ------------------------")
    print(data)
    return data


def run(file_path: str):
    max_page = 0
    if os.path.exists(file_path):
        end_page = 2
        for page in range(1, end_page):
            start_url = f"https://www.gxghj.cn/cles/list-141-{page}.html"
            html = get_content(start_url)
            articles_list, pagex = parse_articles_list(html)
            for item in articles_list[:1]:
                print(item)
                item_html = get_content(item[0])
                water_data = parse_data(item_html)
                # save_data(file_path, water_data)
                data_list.append(water_data)
            if page == max_page:
                break
    else:
        end_page = 999999
        for page in range(1, end_page):
            start_url = f"https://www.gxghj.cn/cles/list-141-{page}.html"
            html = get_content(start_url)
            articles_list, pagex = parse_articles_list(html)
            if 0 < pagex != max_page:
                max_page = pagex
            for item in articles_list:
                print(item)
                item_html = get_content(item[0])
                water_data = parse_data(item_html)
                data_list.append(water_data)
                # water_data["url"] = item[0]
                # save_data(file_path, water_data)
            if page >= max_page:
                break
            # time.sleep(random.uniform(0, 1))
    save_data(file_path)
    print(f"------------------------ 程序运行结束 ------------------------")


if __name__ == '__main__':
    file_path = "柳州航道养护中心水位通告.xlsx"
    run(file_path)
