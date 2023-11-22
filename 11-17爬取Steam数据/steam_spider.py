import json
import os
import time
from io import BytesIO

import pandas as pd
import requests
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from bs4 import BeautifulSoup
from retrying import retry
from openpyxl.drawing.image import Image
import concurrent.futures

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

heads = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"}

heads2 = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "X-Requested-With": "XMLHttpRequest",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}

proxies = {"http": "http://127.0.0.1:4780", "https": "http://127.0.0.1:4780"}
# 白名单方式（需提前设置白名单）
# 隧道域名:端口号
tunnel = "o422.kdltps.com:15818"


# proxies = {
#     "http": "http://%(proxy)s/" % {"proxy": tunnel},
#     "https": "http://%(proxy)s/" % {"proxy": tunnel}
# }


# 自定义比较函数，将百分比字符串转换为浮点数进行比较
def custom_sort(item):
    return float(item['value'].strip('%'))


def init_session():
    req_session = requests.session()
    req_session.headers = heads2
    return req_session


req_session = init_session()


def agecheckset():
    url = "https://store.steampowered.com/agecheckset/app/2050650/"
    data = {"sessionid": req_session.cookies.get("sessionid"), "ageDay": 1, "ageMonth": "January", "ageYear": "1990"}
    resp = req_session.post(url=url, data=data, proxies=proxies)
    if resp.status_code == 200:
        print(resp)


@retry(wait_fixed=1000)
def get_steam_scout(app_id, language: str):
    url = f"https://www.togeproductions.com/SteamScout/a.php?appID={app_id}&language={language}&purchase_type=all"
    resp = requests.get(url=url, headers=heads, proxies=proxies)
    if resp.status_code == 200:
        json_data = resp.json()
        print(url)
        if json_data["query_summary"]:
            return json_data["query_summary"]
    else:
        print(resp.status_code, resp.text)
    return {}


@retry(wait_fixed=2000)
def get_steam_img(app_id) -> list:
    movie_img_list = []
    url = f"https://store.steampowered.com/app/{app_id}"
    resp = req_session.get(url=url, proxies=proxies)
    if resp.status_code == 200:
        soup = BeautifulSoup(resp.content, 'html.parser')
        soup = soup.select("div[id='highlight_strip_scroll']")
        if not soup:
            return movie_img_list
        print(url, soup)
        for child in soup[0].contents:
            if child == "\n" or child.img == None:
                continue
            elif "thumb_movie" in child.get("id"):
                movie_id = child.get("id").split("thumb_movie_")[1]
                movie_url = f"https://cdn.cloudflare.steamstatic.com/steam/apps/{movie_id}/movie_max.mp4?t={int(time.time())}"
                movie_img_list.append(movie_url)
            elif "thumb_screenshot" in child.get("id") and child.img:
                img_url = child.img.get("src")
                img_url = img_url.replace("116x65", "600x338")
                movie_img_list.append(img_url)
            if len(movie_img_list) == 4:
                break
    print(movie_img_list)
    return movie_img_list


def get_portions(app_id):
    global languages
    all_data = {}
    data = [str(app_id)]
    portions = []
    for key in keys:
        while True:
            d = get_steam_scout(app_id, key)
            if d:
                break
        print(d)
        if key == "all":
            all_data = d
            continue
        if "total_reviews" not in d:
            print(app_id, key)
            continue
        elif d["total_reviews"] == 0:
            portion = "0.00%"
        else:
            portion = round(d["total_reviews"] / all_data["total_reviews"] * 100, 2)
            portion = f"{portion}%"
        portions.append({"name": key, "value": portion})
    portions = sorted(portions, key=custom_sort, reverse=True)
    for p in portions[:3]:
        data.append(languages[p["name"]])
        data.append(p["value"])
    url_list = get_steam_img(app_id)
    data.extend(url_list)
    return data


if __name__ == '__main__':
    datas = []
    languages = {
        'all': '所有',
        'arabic': '阿拉伯语',
        'bulgarian': '保加利亚语',
        'schinese': '简体中文',
        'tchinese': '繁体中文',
        'czech': '捷克语',
        'danish': '丹麦语',
        'dutch': '荷兰语',
        'english': '英语',
        'finnish': '芬兰语',
        'french': '法语',
        'german': '德语',
        'greek': '希腊语',
        'hungarian': '匈牙利语',
        'indonesian': '印尼语',
        'italian': '意大利语',
        'japanese': '日语',
        'koreana': '韩语',
        'norwegian': '挪威语',
        'polish': '波兰语',
        'portuguese': '葡萄牙语',
        'brazilian': '巴西葡萄牙语',
        'romanian': '罗马尼亚语',
        'russian': '俄语',
        'spanish': '西班牙语',
        'latam': '拉丁美洲西班牙语',
        'swedish': '瑞典语',
        'thai': '泰语',
        'turkish': '土耳其语',
        'ukrainian': '乌克兰语',
        'vietnamese': '越南语'
    }
    # 读取 Excel 文件
    file_path = './清单表.xlsx'  # 替换为你的文件路径
    column_name = 'SteamID'  # 替换为你想要获取的列名
    df = pd.read_excel(file_path, converters={column_name: int})
    # 获取某一列的数据
    column_data = df[column_name][1:].tolist()
    keys = list(languages.keys())
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        # 提交任务给线程池
        futures = {executor.submit(get_portions, value): value for value in column_data}
        # 等待所有线程执行完毕
        for future in concurrent.futures.as_completed(futures):
            option_value = futures[future]
            try:
                data = future.result()
                datas.append(data)
            except Exception as e:
                print(f"线程发生异常: {e}")
    # 打印列数据
    # print(column_data)
    # 列名
    # 读取JSON文件
    # file_path = './datas_new.json'
    # with open(file_path, 'r', encoding="utf-8") as file:
    #     # 使用json.load()方法加载JSON文件
    #     datas = json.load(file)

    xxx = False
    for i in datas:
        if len(i) == 7:
            url_list = get_steam_img(i[0])
            if not xxx:
                agecheckset()
                xxx = True
                url_list = get_steam_img(i[0])
            if not url_list:
                print(i)
            i.extend(url_list)
    # Excel 文件路径
    excel_file_path = '清单表 - 测试.xlsx'
    # 创建一个 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    # 遍历数据并将其写入Excel文件
    for row_index, row_data in enumerate(datas, start=1):
        for col_index, link in enumerate(row_data, start=1):
            # 如果是图片链接，则下载图片并插入单元格
            if '.jpg' in link or '.png' in link:
                # 处理图片
                response = requests.get(link, heads)
                if response.status_code == 200:
                    # 将图片字节数据保存到 BytesIO 对象
                    img = Image(BytesIO(response.content))
                    ws.add_image(img, f"{get_column_letter(col_index)}{row_index}")
                else:
                    print(f"Failed to download image for link {link}")
            else:
                # 其他情况直接插入数据
                ws.cell(row=row_index, column=col_index, value=link)
    # 保存 Excel 文件
    wb.save(excel_file_path)
